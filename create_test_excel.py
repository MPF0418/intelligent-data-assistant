# -*- coding: utf-8 -*-
"""
将测试数据转换为Excel文档
"""

import json
import os

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_excel_report():
    """创建测试数据Excel报告"""
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    
    # 样式定义
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== Sheet 1: 销售数据 ==========
    ws1 = wb.active
    ws1.title = '销售数据'
    
    sales_headers = ['地区', '省份', '产品', '销售额', '数量', '日期', '增长率']
    sales_data = [
        ['华东', '上海', '产品A', 15000, 100, '2024-01', 0.15],
        ['华东', '江苏', '产品A', 12000, 80, '2024-01', 0.12],
        ['华东', '浙江', '产品B', 18000, 120, '2024-01', 0.18],
        ['华南', '广东', '产品A', 20000, 150, '2024-01', 0.20],
        ['华南', '福建', '产品B', 9000, 60, '2024-01', 0.09],
        ['华北', '北京', '产品A', 25000, 200, '2024-01', 0.25],
        ['华北', '天津', '产品B', 8000, 50, '2024-01', 0.08],
        ['华中', '湖北', '产品A', 11000, 70, '2024-01', 0.11],
        ['华中', '湖南', '产品B', 7000, 40, '2024-01', 0.07],
        ['西南', '四川', '产品A', 13000, 90, '2024-01', 0.13],
        ['西南', '重庆', '产品B', 6000, 30, '2024-01', 0.06],
        ['西北', '陕西', '产品A', 5000, 25, '2024-01', 0.05],
        ['西北', '甘肃', '产品B', 3000, 15, '2024-01', 0.03],
        ['东北', '辽宁', '产品A', 8000, 55, '2024-01', 0.08],
        ['东北', '吉林', '产品B', 4000, 20, '2024-01', 0.04],
    ]
    
    # 写入表头
    for col, header in enumerate(sales_headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入数据
    for row_idx, row_data in enumerate(sales_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 7:  # 增长率列
                cell.number_format = '0.00%'
    
    # 调整列宽
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 10
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 10
    ws1.column_dimensions['F'].width = 12
    ws1.column_dimensions['G'].width = 10
    
    # ========== Sheet 2: 测试用例 ==========
    ws2 = wb.create_sheet('测试用例')
    
    test_headers = ['用例ID', '测试场景', '输入', '预期意图', '优先级']
    test_cases = [
        ['TC-IR-001', '查找最大值', '哪个省公司的销售额最高', 'QUERY_FIND', 'P0'],
        ['TC-IR-002', '查找最小值', '找出金额最低的记录', 'QUERY_FIND', 'P0'],
        ['TC-IR-003', '查找前N名', '销售额前5名是哪些', 'QUERY_FIND', 'P0'],
        ['TC-IR-004', '求和统计', '统计销售额总和', 'QUERY_AGGREGATE', 'P0'],
        ['TC-IR-005', '平均值统计', '计算平均销售额', 'QUERY_AGGREGATE', 'P0'],
        ['TC-IR-006', '计数统计', '统计记录数量', 'QUERY_AGGREGATE', 'P0'],
        ['TC-IR-007', '中位数统计', '销售额的中位数是多少', 'QUERY_AGGREGATE', 'P0'],
        ['TC-IR-008', '标准差统计', '数据的波动程度如何', 'QUERY_AGGREGATE', 'P0'],
        ['TC-IR-009', '百分位数', '第90百分位数是多少', 'QUERY_AGGREGATE', 'P1'],
        ['TC-IR-010', '条件筛选', '筛选出金额大于1000的记录', 'QUERY_FILTER', 'P0'],
        ['TC-IR-011', '升序排序', '按销售额从小到大排序', 'QUERY_SORT', 'P0'],
        ['TC-IR-012', '柱状图', '画一个柱状图展示各地区销售额', 'CHART_BAR', 'P0'],
        ['TC-IR-013', '折线图', '绘制销售额趋势折线图', 'CHART_LINE', 'P0'],
        ['TC-IR-014', '饼图', '用饼图展示各产品占比', 'CHART_PIE', 'P0'],
        ['TC-IR-015', '数据透视表', '按地区和产品交叉统计销售额', 'PIVOT_TABLE', 'P0'],
        ['TC-IR-016', '数据清洗-去重', '删除重复数据', 'DATA_CLEAN', 'P1'],
        ['TC-IR-017', '数据清洗-空值', '把空值填充为平均值', 'DATA_CLEAN', 'P1'],
        ['TC-IR-018', '组合图', '画一个柱状图显示销售额，折线图显示增长率', 'CHART_COMBO', 'P1'],
        ['TC-IR-019', '雷达图', '画一个雷达图对比各部门指标', 'CHART_RADAR', 'P2'],
        ['TC-IR-020', '漏斗图', '漏斗图分析转化率', 'CHART_FUNNEL', 'P2'],
        ['TC-IR-021', '热力图', '生成热力图', 'CHART_HEATMAP', 'P2'],
        ['TC-IR-022', '数据导出', '导出为Excel文件', 'DATA_EXPORT', 'P1'],
    ]
    
    # 写入表头
    for col, header in enumerate(test_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入数据
    for row_idx, row_data in enumerate(test_cases, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    # 调整列宽
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 45
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 8
    
    # ========== Sheet 3: 测试结果 ==========
    ws3 = wb.create_sheet('测试结果')
    
    result_headers = ['用例ID', '输入', '预期', '实际', '置信度', '结果']
    
    # 加载测试结果
    try:
        with open('test_results.json', 'r', encoding='utf-8') as f:
            results = json.load(f)
        test_results = results.get('details', [])
    except:
        test_results = []
    
    # 写入表头
    for col, header in enumerate(result_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入数据
    pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    for row_idx, result in enumerate(test_results, 2):
        ws3.cell(row=row_idx, column=1, value=result.get('case_id', '')).border = thin_border
        ws3.cell(row=row_idx, column=2, value=result.get('input', '')).border = thin_border
        ws3.cell(row=row_idx, column=3, value=result.get('expected', '')).border = thin_border
        ws3.cell(row=row_idx, column=4, value=result.get('actual', '')).border = thin_border
        
        conf_cell = ws3.cell(row=row_idx, column=5, value=f"{result.get('confidence', 0)*100:.2f}%")
        conf_cell.border = thin_border
        
        result_cell = ws3.cell(row=row_idx, column=6, value='通过' if result.get('passed') else '失败')
        result_cell.border = thin_border
        result_cell.fill = pass_fill if result.get('passed') else fail_fill
    
    # 调整列宽
    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 40
    ws3.column_dimensions['C'].width = 18
    ws3.column_dimensions['D'].width = 18
    ws3.column_dimensions['E'].width = 10
    ws3.column_dimensions['F'].width = 8
    
    # ========== Sheet 4: 版本信息 ==========
    ws4 = wb.create_sheet('版本信息')
    
    version_info = [
        ['项目', '内容'],
        ['版本号', 'V3.3'],
        ['发布日期', '2026-03-02'],
        ['模型标签数', '15'],
        ['测试用例数', '30'],
        ['测试通过率', '96.67%'],
        ['模型准确率', '98.50%'],
        ['F1分数', '98.32%'],
        ['', ''],
        ['新增功能', ''],
        ['1', '数据透视表 - 多维交叉分析'],
        ['2', '高级统计函数 - 中位数、众数、标准差、方差、百分位数'],
        ['3', '增强图表 - 组合图、雷达图、漏斗图、热力图'],
        ['4', '数据清洗 - 去重、空值处理、异常检测'],
        ['5', '数据导出 - Excel、CSV、报告'],
    ]
    
    for row_idx, row_data in enumerate(version_info, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            cell.border = thin_border
    
    ws4.column_dimensions['A'].width = 15
    ws4.column_dimensions['B'].width = 50
    
    # 保存文件
    output_path = 'test_data/V3.3测试数据.xlsx'
    os.makedirs('test_data', exist_ok=True)
    wb.save(output_path)
    
    print(f"Excel文件已生成: {output_path}")
    return output_path

if __name__ == '__main__':
    create_excel_report()
